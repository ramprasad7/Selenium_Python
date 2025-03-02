import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

file_path= "C:\\Users\\ram7\\Downloads\\download.xlsx"
fruit_name = "Apple"
new_price = '287'

def update_excel_data(excel_path,search_fruit,column_name,value) -> None:
    book = openpyxl.load_workbook(excel_path)
    sheet = book.active
    Dict: dict = {}
    for i in range(1, sheet.max_column+1):
        if sheet.cell(row=1,column=i).value == column_name:
            Dict["col"] = i

    for i in range(1,sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            if sheet.cell(row=i,column=j).value == search_fruit:
                Dict["row"] = i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value = value
    book.save(file_path)


driver: webdriver = webdriver.Chrome()

driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.maximize_window()
driver.implicitly_wait(4)


driver.find_element(By.CSS_SELECTOR,".button").click()
time.sleep(2)

update_excel_data(file_path,fruit_name,"price",new_price)

file_upload = driver.find_element(By.XPATH,"//input[@type='file']")
file_upload.send_keys(file_path)

wait = WebDriverWait(driver, 5)
upload_success_locator = (By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(upload_success_locator))
#print(driver.find_element(*upload_success_locator).text)


price_col_id = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
#print(price)
price =driver.find_element(By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+price_col_id+"-undefined']").text

print(price)
assert price == new_price




driver.close()